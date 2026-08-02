"""Generate one Holland training dataset per examination year.

The source workbooks are read by column name (not by position), so a changed
ordering of subjects between years cannot silently change the score calculation.
Run from any directory, for example:

    python ai/data/gen_dataset.py --year all
    python ai/data/gen_dataset.py --year 2025
"""

from __future__ import annotations

import argparse
import random
import re
from collections import defaultdict
from pathlib import Path
from typing import Iterable

import pandas as pd
from openpyxl import load_workbook


DATA_DIR = Path(__file__).resolve().parent
YEAR_FILES = {
    "2025": DATA_DIR / "2025" / "diemthi_thpt_2025.xlsx",
    "2026": DATA_DIR / "2026" / "diemthi_thpt_2026.xlsx",
}

SUBJECT_ALIASES = {
    "tieng anh": "Ngoai Ngu",
    "ngoai ngu": "Ngoai Ngu",
}


def canonical_subject(value: object) -> str:
    name = re.sub(r"\s+", " ", str(value).strip())
    return SUBJECT_ALIASES.get(name.casefold(), name)


def load_reference_tables() -> tuple[dict[str, list[str]], list[dict], dict[str, tuple[str, ...]]]:
    """Load the subject combinations, majors and Holland mapping tables."""
    combinations = pd.read_csv(DATA_DIR / "to_hop_mon.csv", encoding="utf-8")
    grouped = combinations.groupby("MaToHop")["MonHoc"].apply(list).to_dict()
    to_hop = {
        code: tuple(dict.fromkeys(canonical_subject(subject) for subject in subjects))
        for code, subjects in grouped.items()
    }

    majors_df = pd.read_excel(DATA_DIR / "ThongKeByDuy.xlsx")
    majors = []
    for _, row in majors_df.iterrows():
        codes = {x.strip() for x in str(row["ToHopXetTuyenChung"]).split(",") if x.strip()}
        majors.append(
            {
                "label": str(row["Label"]).strip(),
                "tohop": codes,
                "diem_min": float(row["DiemMin"]),
            }
        )

    holland_df = pd.read_excel(DATA_DIR / "Holland_ThongKe_Mapping.xlsx")
    holland: dict[str, list[str]] = defaultdict(list)
    for _, row in holland_df.iterrows():
        # Current mapping files use commas; accept semicolons too for older files.
        labels = re.split(r"[;,]", str(row["NhomNganhNghe"]))
        for label in labels:
            label = label.strip()
            if label:
                holland[label].append(str(row["MaTinhCach"]).strip())

    return holland, majors, to_hop


def iter_score_rows(path: Path) -> Iterable[dict[str, float]]:
    """Stream non-empty score rows from a large XLSX without loading it all."""
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook[workbook.sheetnames[0]]
        rows = sheet.iter_rows(values_only=True)
        headers = next(rows)
        header_names = [canonical_subject(x) for x in headers]
        if "SOBAODANH" not in headers:
            raise ValueError(f"{path.name}: thiếu cột SOBAODANH")

        subject_indexes = {
            name: index
            for index, name in enumerate(header_names)
            if index >= 2 and name not in {"STT", "SOBAODANH"}
        }

        for values in rows:
            if not values or values[1] is None:
                continue
            scores = {}
            for subject, index in subject_indexes.items():
                value = values[index] if index < len(values) else None
                if value is not None:
                    try:
                        scores[subject] = float(value)
                    except (TypeError, ValueError):
                        continue
            if scores:
                yield scores
    finally:
        workbook.close()


def generate_sources(sources: list[tuple[str, Path]], output: Path,
                     holland: dict[str, list[str]], majors: list[dict], to_hop: dict,
                     max_sample: int = 10_000, seed: int = 42) -> Path:
    for year, source in sources:
        if not source.exists():
            raise FileNotFoundError(f"Không tìm thấy file điểm {year}: {source}")

    majors_by_combo: dict[str, list[dict]] = defaultdict(list)
    for major in majors:
        for code in major["tohop"]:
            majors_by_combo[code].append(major)

    # Reservoir sampling keeps at most max_sample rows per major while streaming
    # millions of candidate rows from the workbook.
    rng = random.Random(seed)
    buckets: dict[str, list[dict]] = defaultdict(list)
    seen: defaultdict[str, int] = defaultdict(int)
    row_counts: dict[str, int] = {}

    for year, source in sources:
        row_counts[year] = 0
        for scores in iter_score_rows(source):
            row_counts[year] += 1
            for code, subjects in to_hop.items():
                if not all(subject in scores for subject in subjects):
                    continue
                score = round(sum(scores[subject] for subject in subjects), 2)
                for major in majors_by_combo.get(code, []):
                    if score < major["diem_min"]:
                        continue
                    label = major["label"]
                    record = {
                        "MaToHop": code,
                        "DiemToHop": score,
                        "NhomTinhCach": rng.choice(holland.get(label, [""])),
                        "NhomNganh": label,
                    }
                    seen[label] += 1
                    bucket = buckets[label]
                    if len(bucket) < max_sample:
                        bucket.append(record)
                    else:
                        slot = rng.randrange(seen[label])
                        if slot < max_sample:
                            bucket[slot] = record

    records = [record for bucket in buckets.values() for record in bucket]
    rng.shuffle(records)
    output.parent.mkdir(parents=True, exist_ok=True)
    pd.DataFrame(records, columns=["MaToHop", "DiemToHop", "NhomTinhCach", "NhomNganh"]).to_csv(
        output, index=False, encoding="utf-8-sig"
    )

    read_summary = ", ".join(f"{year}: {count:,}" for year, count in row_counts.items())
    print(f"Đã đọc {read_summary} dòng điểm, sinh {len(records):,} dòng -> {output}")
    print(pd.Series([record["NhomNganh"] for record in records]).value_counts().to_string())
    return output


def generate_year(year: str, holland: dict[str, list[str]], majors: list[dict], to_hop: dict,
                  max_sample: int = 10_000, seed: int = 42) -> Path:
    output = DATA_DIR / year / f"dataset_balanced_holland_{year}.csv"
    return generate_sources([(year, YEAR_FILES[year])], output, holland, majors, to_hop,
                            max_sample, seed)


def generate_mixed(holland: dict[str, list[str]], majors: list[dict], to_hop: dict,
                   max_sample: int = 10_000, seed: int = 42) -> Path:
    output = DATA_DIR / "mixed" / "dataset_balanced_holland_mixed.csv"
    return generate_sources(
        [("2025", YEAR_FILES["2025"]), ("2026", YEAR_FILES["2026"])],
        output, holland, majors, to_hop, max_sample, seed
    )


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--year", choices=["2025", "2026", "all", "mixed"], default="all")
    parser.add_argument("--max-sample", type=int, default=10_000)
    args = parser.parse_args()

    holland, majors, to_hop = load_reference_tables()
    if args.year == "mixed":
        generate_mixed(holland, majors, to_hop, args.max_sample)
    else:
        years = ["2025", "2026"] if args.year == "all" else [args.year]
        for year in years:
            generate_year(year, holland, majors, to_hop, args.max_sample)


if __name__ == "__main__":
    main()
